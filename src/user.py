from openpyxl import load_workbook, Workbook
from tksave import *

user_name = 1
user_id = 2
user_money = 3
user_level = 4

default_money = 1000

wb = load_workbook(xlsxDir)
sheet = wb.active # 활성 중인 워크 시트 선택

'''
UserDB가 담겨 있는 xlsx 파일 열기
'''
def loadFile():
    wb = load_workbook(xlsxDir)
    sheet = wb.active 

'''
UserDB가 담겨 있는 xlsx 파일 열기
'''
def saveFile():
    wb.save(xlsxDir)
    wb.close()

# ============================ UserCheck ============================    

'''
User가 몇 명 있는지 세어 줌
'''
def checkUserNum(): 

    loadFile()
    count = 0

    # max_row => 값이 있는 가장 마지막 row 가지고 옴
    for row in range(2, sheet.max_row + 1): 
        if sheet.cell(row, user_name).value != None:
            count = count + 1
        else:
            count = count

    return count

'''
첫 번째로 빈 셀을 탐색 
같은 칸에 덮어씌우는 것을 방지하기 위함
'''
def checkRow():

    loadFile()

    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row, 1).value is None:
            return row
            break
    
    _result = sheet.max_row + 1

    saveFile()

    return _result

'''
유저가 등록되어 있는지 확인
@param _name 유저의 이름
@param _id 유저의 고유 번호
'''
def checkUser(_name, _id):

    loadFile()

    userNum = checkUserNum()

    # 이름과 id 탐색
    for row in range(2, 3 + userNum):
        if sheet.cell(row, user_name).value == _name and sheet.cell(row, user_id).value == hex(_id):
            saveFile()

            return True, row
            break
    
    saveFile()

    return False, None

# ========================== AccountManagement ==========================

'''
유저 회원 가입 시 DB에 추가
@param _name 유저의 이름
@param _id 유저의 고유 번호
'''
def signUp(_name, _id):

    loadFile()

    _row = checkRow()

    sheet.cell(row = _row, column = user_name, value = _name)
    sheet.cell(row = _row, column = user_id, value = hex(_id))
    sheet.cell(row = _row, column = user_level, value = 1)
    sheet.cell(row = _row, column = user_money, value = default_money)

    saveFile()
    
# =============================== Info ===============================

'''
유저 정보에서 level과 money 찾기
@param _row 유저가 현재 있는 행
'''
def userInfo(_row):
    
    loadFile()

    _level = sheet.cell(_row, user_level).value
    _money = sheet.cell(_row, user_money).value

    saveFile()

    return _level, _money

# =============================== Money ===============================

'''
유저 정보에서 돈 정보 찾기
@param _name 유저의 이름
@param _row 유저가 현재 있는 행
'''
def getMoney(_name, _row):

    loadFile()

    result = sheet.cell(_row, user_money).value

    saveFile()

    return result

'''
돈 얼마나 변했는지 (도박 기능 돈 변동 때문에 넣은 함수)
@param _target 상대 유저 (나중에 송금 기능 추가할 때 사용할 것)
@param _row 유저가 현재 있는 행
@param _amount 돈 얼마나 변했는지
'''
def modifyMoney(_target, _row, _amount):

    loadFile()

    sheet.cell(_row, user_money).value += _amount

    saveFile()

'''
돈 얼마나 잃었는지
ㄴ> 내정보에서 잃은 돈 보여 주기 위해 사용
@param _target 상대 유저 (나중에 송금 기능 추가할 때 사용할 것)
@param _row 유저가 현재 있는 행
@param _amount 돈 얼마나 변했는지
'''
def moneyLoss(_target, _row, _amount):

    loadFile()

    sheet.cell(_row, user_loss).value += _amount

    saveFile()